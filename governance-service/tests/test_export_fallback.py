"""F3（2026-07-10，凍結例外已登記於手冊 §1.1）：rule-run Excel 匯出的 _RUN_CACHE 為
process-local——重啟/多 worker 後 cache miss，舊行為對「DB 已有完整結果」的 succeeded run
回 409。本檔鎖定新行為：cache miss 時由 store 重建匯出（200 xlsx）；查無 run 才 404；
未完成 run 仍 409（誠實：不匯出半成品）。
"""
from __future__ import annotations

import importlib

import pytest
from fastapi.testclient import TestClient

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture()
def app_module(tmp_path, monkeypatch):
    monkeypatch.setenv("GOV_DB_PATH", str(tmp_path / "gov.db"))
    import app as module

    importlib.reload(module)
    return module


def _seed_succeeded_run(app_module) -> str:
    from rule_engine.models import RuleResult, RuleRunResult

    run_id = app_module.store.create_run("mv_test", "C:/fake/model.ifc", "default")
    run = RuleRunResult(
        rule_set="default",
        version="1",
        target_summary={"DOOR-FIRERATING-REQUIRED": 2},
        total=2,
        passed=1,
        failed=1,
        errored=0,
        score=50.0,
        unique_elements=2,
        results=[
            RuleResult(
                ifc_guid="2O2Fr$t4X7Zf8NOew3FL9r", ifc_type="IfcDoor", ifc_name="D-01",
                rule_code="DOOR-FIRERATING-REQUIRED", severity="high", status="fail",
                message="missing FireRating", evidence={"pset": "Pset_DoorCommon"},
            ),
        ],
        warnings=[],
    )
    app_module.store.complete_run(run_id, run)
    return run_id


def test_export_rebuilds_from_store_after_cache_miss(app_module):
    run_id = _seed_succeeded_run(app_module)
    assert run_id not in app_module._RUN_CACHE  # 模擬重啟後：cache 一定沒有
    client = TestClient(app_module.app)
    res = client.get(f"/api/rule-runs/{run_id}/export?fmt=excel")
    assert res.status_code == 200, res.text
    assert res.headers["content-type"].startswith(XLSX_MIME)
    assert len(res.content) > 1000  # 真 workbook bytes，非空殼

    # 重建的 run 內容忠於 DB（Failed Elements sheet 含該 guid 列）。
    from io import BytesIO
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(res.content))
    rows = list(wb["Failed Elements"].iter_rows(values_only=True))
    assert any("2O2Fr$t4X7Zf8NOew3FL9r" in (row or ()) for row in rows[1:])


def test_export_404_when_run_absent(app_module):
    client = TestClient(app_module.app)
    res = client.get("/api/rule-runs/rr_nope/export?fmt=excel")
    assert res.status_code == 404


def test_export_409_when_run_not_succeeded(app_module):
    run_id = app_module.store.create_run("mv_test", "C:/fake/model.ifc", "default")  # queued，未完成
    client = TestClient(app_module.app)
    res = client.get(f"/api/rule-runs/{run_id}/export?fmt=excel")
    assert res.status_code == 409
